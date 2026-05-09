import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generar_pdf(datos_ventas: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = styles['Heading1']
    title_style.alignment = 1  # Center
    elements.append(Paragraph("Reporte de Ventas", title_style))
    elements.append(Spacer(1, 12))

    # Period
    periodo = datos_ventas.get('periodo', {})
    elements.append(Paragraph(f"<b>Periodo:</b> {periodo.get('fecha_inicio')} a {periodo.get('fecha_fin')}", styles['Normal']))
    elements.append(Spacer(1, 12))

    # Summary
    resumen = datos_ventas.get('resumen', {})
    elements.append(Paragraph("<b>Resumen Financiero:</b>", styles['Heading3']))
    resumen_data = [
        ["Número de Ventas", str(resumen.get('num_ventas', 0))],
        ["Subtotal", f"${resumen.get('subtotal', 0):.2f}"],
        ["Impuestos", f"${resumen.get('impuestos', 0):.2f}"],
        ["Total Vendido", f"${resumen.get('total_ventas', 0):.2f}"]
    ]
    t_resumen = Table(resumen_data, colWidths=[200, 100])
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.whitesmoke),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0,0), (-1,-1), 1, colors.lightgrey)
    ]))
    elements.append(t_resumen)
    elements.append(Spacer(1, 20))

    # Payment Methods
    elements.append(Paragraph("<b>Desglose por Método de Pago:</b>", styles['Heading3']))
    metodos = datos_ventas.get('metodos_pago', [])
    metodos_data = [["Método", "Ventas", "Total"]]
    for m in metodos:
        metodos_data.append([m['metodo_pago'].capitalize(), str(m['ventas']), f"${m['total']:.2f}"])
    
    t_metodos = Table(metodos_data, colWidths=[150, 100, 100])
    t_metodos.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(t_metodos)
    elements.append(Spacer(1, 20))

    # Top Products
    elements.append(Paragraph("<b>Top Productos Vendidos:</b>", styles['Heading3']))
    productos = datos_ventas.get('productos_vendidos', [])
    prod_data = [["ID", "Nombre", "Unidades", "Importe"]]
    for p in productos:
        prod_data.append([str(p['id_producto']), p['nombre'], str(p['unidades']), f"${p['importe']:.2f}"])

    t_prod = Table(prod_data, colWidths=[50, 200, 80, 100])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.steelblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(t_prod)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generar_excel(datos_ventas: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(bold=True, size=16)

    # Title
    ws['A1'] = "Reporte de Ventas"
    ws['A1'].font = title_font
    
    periodo = datos_ventas.get('periodo', {})
    ws['A2'] = f"Periodo: {periodo.get('fecha_inicio')} a {periodo.get('fecha_fin')}"
    
    row = 4

    # Summary
    ws.cell(row=row, column=1, value="Resumen Financiero").font = Font(bold=True)
    row += 1
    resumen = datos_ventas.get('resumen', {})
    ws.cell(row=row, column=1, value="Número de Ventas")
    ws.cell(row=row, column=2, value=resumen.get('num_ventas', 0))
    row += 1
    ws.cell(row=row, column=1, value="Total Vendido")
    ws.cell(row=row, column=2, value=resumen.get('total_ventas', 0))
    row += 2

    # Payment Methods
    ws.cell(row=row, column=1, value="Desglose por Método de Pago").font = Font(bold=True)
    row += 1
    headers = ["Método", "Ventas", "Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    
    metodos = datos_ventas.get('metodos_pago', [])
    for m in metodos:
        ws.cell(row=row, column=1, value=m['metodo_pago'].capitalize())
        ws.cell(row=row, column=2, value=m['ventas'])
        ws.cell(row=row, column=3, value=m['total'])
        row += 1
    row += 2

    # Top Products
    ws.cell(row=row, column=1, value="Top Productos Vendidos").font = Font(bold=True)
    row += 1
    headers_prod = ["ID", "Nombre", "Unidades", "Importe"]
    for col, h in enumerate(headers_prod, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1

    productos = datos_ventas.get('productos_vendidos', [])
    for p in productos:
        ws.cell(row=row, column=1, value=p['id_producto'])
        ws.cell(row=row, column=2, value=p['nombre'])
        ws.cell(row=row, column=3, value=p['unidades'])
        ws.cell(row=row, column=4, value=p['importe'])
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
